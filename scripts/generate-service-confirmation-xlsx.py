"""
產生給老闆填的「服務項目確認表」Excel 檔。

執行：python3 scripts/generate-service-confirmation-xlsx.py
輸出：docs/服務項目確認表-給老闆-2026-04-24.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_PATH = "/Users/ryan/Documents/VS_code/理髮廳/docs/服務項目確認表-給老闆-2026-04-24.xlsx"

# 配色（淡色系，列印和螢幕都舒適）
FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_SECTION = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
FILL_CONFIRMED = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_FILLIN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_CONSULT = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

FONT_HEADER = Font(name="PingFang TC", size=14, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="PingFang TC", size=12, bold=True, color="1F4E79")
FONT_BODY = Font(name="PingFang TC", size=11)
FONT_BODY_BOLD = Font(name="PingFang TC", size=11, bold=True)
FONT_NOTE = Font(name="PingFang TC", size=10, italic=True, color="595959")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = WRAP_CENTER
    ws.row_dimensions[row].height = 28


def write_section(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = WRAP_LEFT
    ws.row_dimensions[row].height = 24


def write_note(ws, row, text, span):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_NOTE
    cell.alignment = WRAP_LEFT
    ws.row_dimensions[row].height = 30


def write_row(ws, row, values, fills=None, fonts=None, heights=24):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border = BORDER
        cell.alignment = WRAP_CENTER if isinstance(v, (int, float)) else WRAP_LEFT
        cell.font = (fonts[i - 1] if fonts and fonts[i - 1] else FONT_BODY)
        if fills and fills[i - 1]:
            cell.fill = fills[i - 1]
    ws.row_dimensions[row].height = heights


def build():
    wb = Workbook()

    # ============================================================
    # Sheet 1: 服務項目主表
    # ============================================================
    ws = wb.active
    ws.title = "服務項目"
    set_column_widths(ws, [10, 18, 12, 14, 14, 14, 14, 30])

    row = 1
    write_header(ws, row, "1008 Hair Studio — 服務項目確認表", 8)
    row += 1

    write_note(
        ws, row,
        "Ken 老師您好！麻煩您檢視下表，黃底的欄位是請您協助填寫的地方，"
        "綠底是我已確認的內容請 double check，橘底是走「諮詢流程」不開給客人自己勾選的項目。"
        "填完回傳給 Ryan 就可以了，任何不確定的地方先空著沒關係。",
        8,
    )
    row += 1

    # 欄位標題
    headers = ["類別", "項目名稱", "時長(小時)", "價格(NT$)", "是否含洗",
               "綁護髮", "可否單獨預約", "備註／請補充"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = WRAP_CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    # -------- 剪髮類 --------
    write_section(ws, row, "剪髮類（我理解只有三種：男生 / 女生 / 學童，含洗+剪）", 8)
    row += 1

    cut_rows = [
        ("剪髮", "男生（洗+剪）", 1, 1000, "是", "不適用", "可以", ""),
        ("剪髮", "女生（洗+剪）", 1, "", "是", "不適用", "可以", ""),
        ("剪髮", "學童（洗+剪）", 1, "", "是", "不適用", "可以", ""),
        ("剪髮", "（如有其他剪髮項目請加列）", "", "", "", "", "", ""),
    ]
    for data in cut_rows:
        fills = [FILL_CONFIRMED, FILL_CONFIRMED,
                 FILL_CONFIRMED if data[2] else FILL_FILLIN,
                 FILL_CONFIRMED if data[3] else FILL_FILLIN,
                 FILL_CONFIRMED if data[4] else FILL_FILLIN,
                 FILL_CONFIRMED if data[5] else FILL_FILLIN,
                 FILL_CONFIRMED if data[6] else FILL_FILLIN,
                 FILL_FILLIN]
        write_row(ws, row, data, fills)
        row += 1

    # -------- 燙髮 --------
    write_section(ws, row, "燙髮類", 8)
    row += 1
    perm_rows = [
        ("燙髮", "燙髮（標準 4 小時）", 4, "", "是", "", "可以", "護髮是不是一定綁定？"),
        ("燙髮", "（如分溫塑/縮毛矯正/一般燙請加列）", "", "", "", "", "", ""),
    ]
    for data in perm_rows:
        fills = [FILL_CONFIRMED, FILL_CONFIRMED,
                 FILL_CONFIRMED if data[2] else FILL_FILLIN,
                 FILL_FILLIN,
                 FILL_CONFIRMED if data[4] else FILL_FILLIN,
                 FILL_FILLIN, FILL_CONFIRMED if data[6] else FILL_FILLIN,
                 FILL_FILLIN]
        write_row(ws, row, data, fills)
        row += 1

    # -------- 染髮 --------
    write_section(ws, row, "染髮類（我記得您說「染必綁護」— 請再確認）", 8)
    row += 1
    dye_rows = [
        ("染髮", "補染（髮根）", 2, "", "是", "是", "可以", ""),
        ("染髮", "全頭染", 3, "", "是", "是", "可以", "原本說 2.5–3 小時一律算 3"),
    ]
    for data in dye_rows:
        fills = [FILL_CONFIRMED, FILL_CONFIRMED,
                 FILL_CONFIRMED, FILL_FILLIN,
                 FILL_CONFIRMED, FILL_CONFIRMED, FILL_CONFIRMED, FILL_CONFIRMED]
        write_row(ws, row, data, fills)
        row += 1

    # -------- 漂髮（諮詢流程）--------
    write_section(ws, row, "漂髮類（走「諮詢流程」，不給客人自己在 LINE 勾選）", 8)
    row += 1
    bleach_rows = [
        ("漂髮", "純漂（退乾淨色）", "3 小時/次 × 至少 3 次", "", "是", "視情況", "不可（諮詢）", ""),
        ("漂髮", "漂+染（調色）", "6 小時起跳", "", "是", "是", "不可（諮詢）", ""),
    ]
    for data in bleach_rows:
        fills = [FILL_CONSULT] * 8
        # 價格欄是需要填的
        fills[3] = FILL_FILLIN
        write_row(ws, row, data, fills)
        row += 1

    # -------- 附加項目 --------
    write_section(ws, row, "附加項目（時間/價格/能否單獨預約都請您幫忙確認）", 8)
    row += 1
    add_rows = [
        ("附加", "瀏海修", "", "", "", "", "", ""),
        ("附加", "護髮（獨立，非染燙附加）", "", "", "", "", "", ""),
        ("附加", "西髮", "", "", "", "", "", ""),
    ]
    for data in add_rows:
        fills = [FILL_CONFIRMED, FILL_CONFIRMED,
                 FILL_FILLIN, FILL_FILLIN, FILL_FILLIN, FILL_FILLIN, FILL_FILLIN, FILL_FILLIN]
        write_row(ws, row, data, fills)
        row += 1

    # -------- 非預約類 --------
    write_section(ws, row, "非預約類（不進預約系統，但會在營收報表手動輸入金額）", 8)
    row += 1
    other_rows = [
        ("商品", "洗髮產品 / 造型品等", "—", "依商品", "不適用", "不適用", "—", "老闆結帳時手動輸入金額"),
        ("其他", "廠商進貨", "—", "", "不適用", "不適用", "—", ""),
        ("其他", "瀏海後修、回整", "—", "", "不適用", "不適用", "—", ""),
    ]
    for data in other_rows:
        fills = [FILL_CONFIRMED] * 8
        fills[3] = FILL_FILLIN  # 廠商價格需填
        write_row(ws, row, data, fills)
        row += 1

    # 圖例
    row += 2
    write_section(ws, row, "顏色圖例", 8)
    row += 1

    legend_items = [
        ("綠底", "已確認 — 請 double check 有沒有錯", FILL_CONFIRMED),
        ("黃底", "請老闆協助填寫", FILL_FILLIN),
        ("橘底", "走諮詢流程 — 不開給客人在 LINE 自己勾選", FILL_CONSULT),
    ]
    for color_name, desc, fill in legend_items:
        ws.cell(row=row, column=1, value=color_name).fill = fill
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=1).alignment = WRAP_CENTER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
        c = ws.cell(row=row, column=2, value=desc)
        c.alignment = WRAP_LEFT
        c.font = FONT_BODY
        c.border = BORDER
        row += 1

    # ============================================================
    # Sheet 2: 基本營運設定
    # ============================================================
    ws2 = wb.create_sheet("基本營運設定")
    set_column_widths(ws2, [28, 36, 30])

    write_header(ws2, 1, "基本營運設定確認", 3)

    ws2.cell(row=2, column=1, value="項目").font = FONT_BODY_BOLD
    ws2.cell(row=2, column=1).fill = FILL_HEADER
    ws2.cell(row=2, column=1).font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
    ws2.cell(row=2, column=1).alignment = WRAP_CENTER
    ws2.cell(row=2, column=1).border = BORDER
    ws2.cell(row=2, column=2, value="我理解的內容").font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
    ws2.cell(row=2, column=2).fill = FILL_HEADER
    ws2.cell(row=2, column=2).alignment = WRAP_CENTER
    ws2.cell(row=2, column=2).border = BORDER
    ws2.cell(row=2, column=3, value="確認 / 要修正").font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
    ws2.cell(row=2, column=3).fill = FILL_HEADER
    ws2.cell(row=2, column=3).alignment = WRAP_CENTER
    ws2.cell(row=2, column=3).border = BORDER
    ws2.row_dimensions[2].height = 26

    settings = [
        ("公休日", "禮拜三、禮拜日"),
        ("營業時間", "每天 11:00 – 20:00"),
        ("時段單位", "1 小時一格（不做半小時）"),
        ("預約最遠可預約多久", "45 天（1.5 個月）內"),
        ("超過 45 天的客人", "走「諮詢管道」，由老闆親自回覆確認"),
        ("每筆預約的確認方式", "老闆一打開手機會跳出確認視窗，按「已知道」才進行事曆"),
    ]
    r = 3
    for label, content in settings:
        ws2.cell(row=r, column=1, value=label).fill = FILL_CONFIRMED
        ws2.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws2.cell(row=r, column=1).alignment = WRAP_LEFT
        ws2.cell(row=r, column=1).border = BORDER
        ws2.cell(row=r, column=2, value=content).fill = FILL_CONFIRMED
        ws2.cell(row=r, column=2).font = FONT_BODY
        ws2.cell(row=r, column=2).alignment = WRAP_LEFT
        ws2.cell(row=r, column=2).border = BORDER
        ws2.cell(row=r, column=3, value="").fill = FILL_FILLIN
        ws2.cell(row=r, column=3).alignment = WRAP_LEFT
        ws2.cell(row=r, column=3).border = BORDER
        ws2.row_dimensions[r].height = 36
        r += 1

    # ============================================================
    # Sheet 3: 綁定 / 諮詢 / 付款流程確認
    # ============================================================
    ws3 = wb.create_sheet("規則與流程")
    set_column_widths(ws3, [28, 48, 30])

    write_header(ws3, 1, "規則與流程確認", 3)

    for i, h in enumerate(["問題", "選項", "請勾選 / 補充"], start=1):
        cell = ws3.cell(row=2, column=i, value=h)
        cell.font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = WRAP_CENTER
        cell.border = BORDER
    ws3.row_dimensions[2].height = 26

    rules = [
        ("染髮綁護", "染的時候一定要綁護髮嗎？", "(a) 對，直接算進染的價格  (b) 可分開，護是選配"),
        ("燙髮綁護", "燙的時候一定要綁護髮嗎？", "(a) 對，直接算進燙的價格  (b) 可分開，護是選配"),
        ("不洗只剪", "要不要保留「不洗只剪」的便宜選項？", "(a) 要保留，價格：______元  (b) 不要，統一含洗"),
        ("漂髮諮詢流程", "客人點「漂髮諮詢」後，您希望在哪裡回？",
         "(a) LINE 直接回  (b) App 內開「待回覆諮詢」清單  (c) 兩個都要"),
        ("45 天外諮詢", "客人想約 45 天以後（例如國外客），要給他？",
         "(a) 填諮詢表單等回覆  (b) 完全不能按，要打電話"),
        ("商品銷售記錄", "顧客買商品時，您希望？",
         "(a) 結帳時在 App 手動輸入金額  (b) 每晚一起記  (c) 其他："),
        ("付款方式開放", "您希望先開放哪幾種付款方式？",
         "(a) 現金 + 銀行轉帳  (b) 加綠界信用卡（1% 手續費）  (c) 只收現金"),
    ]

    r = 3
    for topic, q, opts in rules:
        ws3.cell(row=r, column=1, value=topic).fill = FILL_CONFIRMED
        ws3.cell(row=r, column=1).font = FONT_BODY_BOLD
        ws3.cell(row=r, column=1).alignment = WRAP_LEFT
        ws3.cell(row=r, column=1).border = BORDER
        ws3.cell(row=r, column=2, value=q + "\n\n" + opts).fill = FILL_CONFIRMED
        ws3.cell(row=r, column=2).font = FONT_BODY
        ws3.cell(row=r, column=2).alignment = WRAP_LEFT
        ws3.cell(row=r, column=2).border = BORDER
        ws3.cell(row=r, column=3, value="").fill = FILL_FILLIN
        ws3.cell(row=r, column=3).alignment = WRAP_LEFT
        ws3.cell(row=r, column=3).border = BORDER
        ws3.row_dimensions[r].height = 62
        r += 1

    # ============================================================
    # Sheet 4: 補充問題 / 自由填寫
    # ============================================================
    ws4 = wb.create_sheet("補充問題")
    set_column_widths(ws4, [48, 48])

    write_header(ws4, 1, "補充問題 — 請老闆簡單回答（文字 / 語音都可）", 2)

    for i, h in enumerate(["問題", "回答"], start=1):
        cell = ws4.cell(row=2, column=i, value=h)
        cell.font = Font(name="PingFang TC", size=11, bold=True, color="FFFFFF")
        cell.fill = FILL_HEADER
        cell.alignment = WRAP_CENTER
        cell.border = BORDER
    ws4.row_dimensions[2].height = 26

    questions = [
        "目前 Excel「新男剪 / 男剪」→ 前面有「新」就是這間店面的新客，沒有就是舊客人，對嗎？",
        "一週中哪幾天 / 哪幾個時段最忙？（會影響系統推薦時段的邏輯）",
        "除了上面列的，還有沒有您常做但 Excel 沒有的項目？",
        "您一個月大概有幾個「國外客 / 特殊預約」要走諮詢流程？",
        "9 折券（剪完 30 天內回來 9 折）這方案您想先試 1 個月看看嗎？",
        "每月營收 / 支出報表，您希望看「週」、「月」、「季」哪個時間單位？",
    ]

    r = 3
    for q in questions:
        ws4.cell(row=r, column=1, value=q).fill = FILL_CONFIRMED
        ws4.cell(row=r, column=1).font = FONT_BODY
        ws4.cell(row=r, column=1).alignment = WRAP_LEFT
        ws4.cell(row=r, column=1).border = BORDER
        ws4.cell(row=r, column=2, value="").fill = FILL_FILLIN
        ws4.cell(row=r, column=2).alignment = WRAP_LEFT
        ws4.cell(row=r, column=2).border = BORDER
        ws4.row_dimensions[r].height = 72
        r += 1

    # 凍結首列
    for sheet in [ws, ws2, ws3, ws4]:
        sheet.freeze_panes = "A3"

    wb.save(OUT_PATH)
    print(f"✅ 已產生：{OUT_PATH}")


if __name__ == "__main__":
    build()
